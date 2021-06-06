import openpyxl
# This class contains all the methods used for Excel Manipulations


# Method for getting the number of Rows in the input excel
def getNumberOfRows(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

# Method for getting the number of Columns in the input excel
def getNumberOfColumns(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

# Method for getting the data from the excel based on row and column
def readData(file, sheetName, rownNumber, columnNumber):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownNumber, column=columnNumber).value
